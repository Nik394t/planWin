import asyncio
import json
from typing import Any

from aiogram import Bot, Dispatcher
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery, Message, FSInputFile
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from config import load_config
from db import (
    create_period_summary,
    create_recurring_task,
    create_task,
    delete_recurring_task,
    delete_task,
    ensure_prayer_rows,
    get_progress,
    get_prayer_entry,
    get_task,
    get_user,
    init_db,
    is_period_finalized,
    is_period_initialized,
    list_history,
    list_prayer_plan,
    list_recurring_tasks,
    list_tasks,
    list_unfinalized_periods,
    list_users,
    log_history,
    mark_period_initialized,
    set_defaults_seeded,
    set_last_ui_message_id,
    set_prayer_entry,
    toggle_task_done,
    update_notify_settings,
    update_recurring_task,
    update_task,
    update_task_lock,
    update_task_source,
    upsert_user,
)
from ui import (
    build_admin_view,
    build_confirm_prompt,
    build_help_view,
    build_history_view,
    build_lock_select_view,
    build_main_menu,
    build_plan_view,
    build_prayer_view,
    build_settings_view,
    build_simple_prompt,
    build_task_detail_view,
    build_task_select_view,
)
from utils import (
    WEEKDAYS_RU,
    context_date_from_period,
    edit_policy,
    is_current_period,
    is_period_over,
    iso_now,
    normalize_time_text,
    parse_title_description,
    period_context_for_date,
    today_date,
    truncate,
)

DEFAULT_DAILY_TASKS = [
    "Почитать Библию",
    "Принять причастие",
]
DEFAULT_PRAYER_TITLE = "Помолиться"

SCOPE_LABELS = {
    "day": "День",
    "week": "Неделя",
    "month": "Месяц",
    "year": "Год",
}


class InputStates(StatesGroup):
    add_task = State()
    edit_task = State()
    set_notify_time = State()
    edit_prayer_title = State()
    edit_prayer_desc = State()


async def ensure_user(db_path: str, message: Message | CallbackQuery) -> dict[str, Any]:
    user = message.from_user
    user_data = {
        "user_id": user.id,
        "username": user.username,
        "first_name": user.first_name,
        "last_name": user.last_name,
    }
    db_user = await upsert_user(db_path, user_data)
    if not db_user.get("defaults_seeded"):
        existing = await list_recurring_tasks(db_path, user.id, "day")
        if not existing:
            for title in DEFAULT_DAILY_TASKS:
                await create_recurring_task(db_path, user.id, "day", title, None)
        await set_defaults_seeded(db_path, user.id, 1)
        db_user = await get_user(db_path, user.id) or db_user
    await ensure_prayer_rows(db_path, user.id)
    return db_user


async def edit_or_send(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    text: str,
    reply_markup,
    message_id: int | None = None,
    force_send: bool = False,
) -> None:
    used_message_id = None if force_send else message_id

    if used_message_id is None and not force_send:
        user = await get_user(db_path, user_id)
        if user and user.get("last_ui_message_id"):
            used_message_id = user["last_ui_message_id"]

    if used_message_id:
        try:
            await bot.edit_message_text(
                text=text,
                chat_id=chat_id,
                message_id=used_message_id,
                reply_markup=reply_markup,
            )
            await set_last_ui_message_id(db_path, user_id, used_message_id)
            return
        except TelegramBadRequest as exc:
            if "message is not modified" in str(exc):
                return
            pass

    sent = await bot.send_message(chat_id=chat_id, text=text, reply_markup=reply_markup)
    await set_last_ui_message_id(db_path, user_id, sent.message_id)


async def show_main_menu(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    is_admin: bool,
    message_id: int | None = None,
) -> None:
    text, kb = build_main_menu(is_admin)
    await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)


async def init_period_if_needed(db_path: str, user_id: int, scope: str, period_key: str) -> None:
    if await is_period_initialized(db_path, user_id, scope, period_key):
        return

    existing = await list_tasks(db_path, user_id, scope, period_key)
    recurring = await list_recurring_tasks(db_path, user_id, scope)
    existing_recurring_ids = {task.get("recurring_id") for task in existing if task.get("recurring_id")}
    existing_by_title: dict[str, dict[str, Any]] = {}
    for task in existing:
        title_key = (task.get("title") or "").strip().lower()
        if title_key and title_key not in existing_by_title:
            existing_by_title[title_key] = task
    for item in recurring:
        if item["id"] in existing_recurring_ids:
            continue
        title_key = (item.get("title") or "").strip().lower()
        candidate = existing_by_title.get(title_key)
        if candidate:
            await update_task_lock(db_path, candidate["id"], 1, item["id"])
            await update_task_source(db_path, candidate["id"], "recurring")
            continue
        await create_task(
            db_path,
            user_id,
            scope,
            period_key,
            item["title"],
            item.get("description"),
            "recurring",
            is_locked=1,
            recurring_id=item["id"],
        )

    await mark_period_initialized(db_path, user_id, scope, period_key)


async def ensure_day_prayer_task(db_path: str, user_id: int, date_key: str) -> None:
    tasks = await list_tasks(db_path, user_id, "day", date_key)
    prayer_task = next(
        (task for task in tasks if task.get("source") in ("prayer", "prayer_default")), None
    )
    await ensure_prayer_rows(db_path, user_id)
    weekday = context_date_from_period("day", date_key).weekday()
    prayer_entry = await get_prayer_entry(db_path, user_id, weekday)

    if prayer_entry and prayer_entry.get("title"):
        desired_title = prayer_entry["title"]
        desired_description = prayer_entry.get("description")
        source = "prayer"
    else:
        desired_title = DEFAULT_PRAYER_TITLE
        desired_description = None
        source = "prayer_default"

    if prayer_task:
        if (
            prayer_task.get("title") != desired_title
            or (prayer_task.get("description") or None) != desired_description
            or prayer_task.get("source") != source
        ):
            await update_task(db_path, prayer_task["id"], desired_title, desired_description)
            await update_task_source(db_path, prayer_task["id"], source)
        if not prayer_task.get("is_locked"):
            await update_task_lock(db_path, prayer_task["id"], 1, None)
    else:
        await create_task(
            db_path,
            user_id,
            "day",
            date_key,
            desired_title,
            desired_description,
            source,
            is_locked=1,
        )


async def finalize_period(db_path: str, user_id: int, scope: str, period_key: str) -> None:
    if await is_period_finalized(db_path, user_id, scope, period_key):
        return
    tasks = await list_tasks(db_path, user_id, scope, period_key)
    if not tasks:
        return
    done_titles = [task.get("title", "") for task in tasks if task.get("is_done")]
    undone_titles = [task.get("title", "") for task in tasks if not task.get("is_done")]
    done_count = len(done_titles)
    total_count = len(tasks)
    await create_period_summary(
        db_path,
        user_id,
        scope,
        period_key,
        done_count,
        total_count,
        done_titles,
        undone_titles,
    )
    await log_history(
        db_path,
        user_id,
        "period_summary",
        {
            "scope": scope,
            "period_key": period_key,
            "done_titles": done_titles,
            "undone_titles": undone_titles,
            "done_count": done_count,
            "total_count": total_count,
        },
    )


async def finalize_overdue_periods(db_path: str, user_id: int) -> None:
    today = today_date()
    for scope in ("day", "week", "month", "year"):
        period_keys = await list_unfinalized_periods(db_path, user_id, scope)
        for period_key in period_keys:
            if is_period_over(scope, period_key, today):
                await finalize_period(db_path, user_id, scope, period_key)


async def show_plan(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    scope: str,
    period_key: str,
    message_id: int | None = None,
    header_text: str | None = None,
    force_send: bool = False,
) -> None:
    await finalize_overdue_periods(db_path, user_id)
    if is_current_period(scope, period_key, today_date()):
        await init_period_if_needed(db_path, user_id, scope, period_key)
        if scope == "day":
            await ensure_day_prayer_task(db_path, user_id, period_key)

    tasks = await list_tasks(db_path, user_id, scope, period_key)

    context_date = context_date_from_period(scope, period_key)
    context = period_context_for_date(context_date)
    tab_context = period_context_for_date(today_date())

    progress = {
        "day": await get_progress(db_path, user_id, "day", context.day_key),
        "week": await get_progress(db_path, user_id, "week", context.week_key),
        "month": await get_progress(db_path, user_id, "month", context.month_key),
        "year": await get_progress(db_path, user_id, "year", context.year_key),
    }

    text, kb = build_plan_view(
        scope,
        period_key,
        context,
        tab_context,
        tasks,
        progress,
        header_text=header_text,
    )
    await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id, force_send=force_send)


async def show_prayer_plan(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    message_id: int | None = None,
) -> None:
    await ensure_prayer_rows(db_path, user_id)
    entries = await list_prayer_plan(db_path, user_id)
    text, kb = build_prayer_view(entries)
    await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)


async def show_settings(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    message_id: int | None = None,
) -> None:
    user = await get_user(db_path, user_id)
    if not user:
        return
    text, kb = build_settings_view(user)
    await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)


def _progress_short(done: int, total: int) -> str:
    if total <= 0:
        return "0/0"
    percent = int(round((done / total) * 100))
    return f"{done}/{total} {percent}%"


async def show_admin_view(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    message_id: int | None = None,
) -> None:
    users = await list_users(db_path)
    enriched = []
    today = today_date()
    context = period_context_for_date(today)

    for u in users:
        done_day, total_day = await get_progress(db_path, u["user_id"], "day", context.day_key)
        done_week, total_week = await get_progress(db_path, u["user_id"], "week", context.week_key)
        done_month, total_month = await get_progress(db_path, u["user_id"], "month", context.month_key)
        done_year, total_year = await get_progress(db_path, u["user_id"], "year", context.year_key)

        display = u.get("username") or ""
        if display:
            display = f"@{display}"
        else:
            display = u.get("first_name") or str(u.get("user_id"))

        enriched.append(
            {
                "user_id": u["user_id"],
                "display": f"{display} ({u['user_id']})",
                "progress_day": _progress_short(done_day, total_day),
                "progress_week": _progress_short(done_week, total_week),
                "progress_month": _progress_short(done_month, total_month),
                "progress_year": _progress_short(done_year, total_year),
            }
        )

    text, kb = build_admin_view(enriched)
    await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)


async def show_history(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    message_id: int | None = None,
) -> None:
    await finalize_overdue_periods(db_path, user_id)
    entries = await list_history(db_path, user_id, limit=15)
    lines = []
    for entry in entries:
        ts = entry.get("ts")
        action = entry.get("action")
        payload = entry.get("payload") or "{}"
        try:
            payload_data = json.loads(payload)
        except json.JSONDecodeError:
            payload_data = {}

        if action == "period_summary":
            scope = payload_data.get("scope", "")
            period_label = payload_data.get("period_key", "")
            scope_label = SCOPE_LABELS.get(scope, scope)
            done_titles = payload_data.get("done_titles") if isinstance(payload_data.get("done_titles"), list) else []
            undone_titles = payload_data.get("undone_titles") if isinstance(payload_data.get("undone_titles"), list) else []
            done_count = payload_data.get("done_count", len(done_titles))
            total_count = payload_data.get("total_count", len(done_titles) + len(undone_titles))
            lines.append(f"{ts} | Итог {scope_label} {period_label}: {done_count}/{total_count}")
            if done_titles:
                lines.append(f"    + {truncate(', '.join(done_titles), 120)}")
            if undone_titles:
                lines.append(f"    - {truncate(', '.join(undone_titles), 120)}")
        elif action == "prayer_update":
            weekday_idx = payload_data.get("weekday")
            weekday_label = WEEKDAYS_RU[weekday_idx] if isinstance(weekday_idx, int) and 0 <= weekday_idx < 7 else ""
            lines.append(f"{ts} | Обновлен молитвенный план: {weekday_label}")
        elif action == "settings_update":
            lines.append(f"{ts} | Настройки уведомлений обновлены")
        elif action == "export":
            lines.append(f"{ts} | Экспорт данных")
        else:
            lines.append(f"{ts} | {action}")

    text, kb = build_history_view(lines)
    await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)


async def export_to_excel(db_path: str, user_id: int) -> str:
    from openpyxl import Workbook
    import aiosqlite
    import os

    wb = Workbook()
    ws_tasks = wb.active
    ws_tasks.title = "tasks"
    ws_tasks.append(
        [
            "id",
            "scope",
            "period_key",
            "title",
            "description",
            "is_done",
            "is_locked",
            "recurring_id",
            "source",
            "created_at",
            "updated_at",
            "done_at",
        ]
    )

    async with aiosqlite.connect(db_path) as db:
        cursor = await db.execute(
            """
            SELECT id, scope, period_key, title, description, is_done, is_locked, recurring_id, source, created_at, updated_at, done_at
            FROM tasks
            WHERE user_id = ?
            ORDER BY id ASC
            """,
            (user_id,),
        )
        tasks = await cursor.fetchall()
        await cursor.close()
        for row in tasks:
            ws_tasks.append(list(row))

        ws_history = wb.create_sheet("history")
        ws_history.append(["id", "ts", "action", "payload"])
        cursor = await db.execute(
            "SELECT id, ts, action, payload FROM history WHERE user_id = ? ORDER BY id ASC",
            (user_id,),
        )
        history_rows = await cursor.fetchall()
        await cursor.close()
        for row in history_rows:
            ws_history.append(list(row))

        ws_prayer = wb.create_sheet("prayer_plan")
        ws_prayer.append(["weekday", "title", "description", "updated_at"])
        cursor = await db.execute(
            "SELECT weekday, title, description, updated_at FROM prayer_plan WHERE user_id = ? ORDER BY weekday ASC",
            (user_id,),
        )
        prayer_rows = await cursor.fetchall()
        await cursor.close()
        for row in prayer_rows:
            ws_prayer.append(list(row))

        ws_summary = wb.create_sheet("period_summaries")
        ws_summary.append(
            [
                "scope",
                "period_key",
                "finalized_at",
                "done_count",
                "total_count",
                "done_titles",
                "undone_titles",
            ]
        )
        cursor = await db.execute(
            """
            SELECT scope, period_key, finalized_at, done_count, total_count, done_titles, undone_titles
            FROM period_summaries
            WHERE user_id = ?
            ORDER BY finalized_at ASC
            """,
            (user_id,),
        )
        summary_rows = await cursor.fetchall()
        await cursor.close()
        for row in summary_rows:
            ws_summary.append(list(row))

    os.makedirs("exports", exist_ok=True)
    filename = f"exports/plan_export_{user_id}_{iso_now().replace(':', '-')}.xlsx"
    wb.save(filename)
    return filename


async def schedule_user_notifications(
    scheduler: AsyncIOScheduler,
    bot: Bot,
    db_path: str,
    user_id: int,
    notify_time: str,
    notify_enabled: int,
) -> None:
    job_id = f"notify_{user_id}"
    if scheduler.get_job(job_id):
        scheduler.remove_job(job_id)

    if not notify_enabled:
        return

    hour, minute = [int(x) for x in notify_time.split(":")]
    scheduler.add_job(
        send_daily_notification,
        CronTrigger(hour=hour, minute=minute),
        id=job_id,
        replace_existing=True,
        args=[bot, db_path, user_id],
    )


async def schedule_all_users(scheduler: AsyncIOScheduler, bot: Bot, db_path: str) -> None:
    users = await list_users(db_path)
    for user in users:
        await schedule_user_notifications(
            scheduler,
            bot,
            db_path,
            user["user_id"],
            user.get("notify_time") or "06:00",
            int(user.get("notify_enabled", 1)),
        )


async def send_daily_notification(bot: Bot, db_path: str, user_id: int) -> None:
    today = today_date()
    reminder_text = (
        "☀️ Доброе утро!\n"
        "Зайди в бота, сформируй план на день, помолись и почитай Библию.\n"
        "Пусть духовный текст ободрит и наставит тебя сегодня. 🙏📖"
    )
    await show_plan(
        bot,
        db_path,
        user_id,
        user_id,
        "day",
        today.strftime("%Y-%m-%d"),
        header_text=reminder_text,
        force_send=True,
    )


def parse_callback(data: str) -> tuple[str, list[str]]:
    parts = data.split(":")
    return parts[0], parts[1:]


def edit_denied_message(scope: str) -> str:
    if scope == "day":
        return "Редактировать дневной план можно только сегодня."
    if scope == "week":
        return "Редактировать недельный план можно только в понедельник текущей недели."
    if scope == "month":
        return "Редактировать месячный план можно только 1-2 числа текущего месяца."
    if scope == "year":
        return "Редактировать годовой план можно только в январе."
    return "Редактирование недоступно."


def edit_confirm_message(scope: str) -> str:
    if scope == "year":
        return "Подтвердите корректировку годового плана в январе."
    return "Подтвердите изменение плана."


async def ensure_edit_allowed(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    scope: str,
    period_key: str,
    action: str,
    message_id: int | None,
    force: bool = False,
    confirm_data: str | None = None,
    back_data: str | None = None,
) -> bool:
    if await is_period_finalized(db_path, user_id, scope, period_key):
        prompt_text, kb = build_simple_prompt(
            "Период уже закрыт. Правки недоступны.", back_data or f"v:{scope}:{period_key}"
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)
        return False

    policy = edit_policy(scope, period_key, today_date())
    if policy == "allow":
        return True
    if policy == "confirm" and force:
        return True
    if policy == "confirm":
        confirm_data = confirm_data or f"cf:{action}:{scope}:{period_key}"
        cancel_data = back_data or f"v:{scope}:{period_key}"
        prompt_text, kb = build_confirm_prompt(edit_confirm_message(scope), confirm_data, cancel_data)
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)
        return False

    prompt_text, kb = build_simple_prompt(edit_denied_message(scope), back_data or f"v:{scope}:{period_key}")
    await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)
    return False


async def ensure_add_allowed(
    bot: Bot,
    db_path: str,
    user_id: int,
    chat_id: int,
    scope: str,
    period_key: str,
    message_id: int | None,
) -> bool:
    if await is_period_finalized(db_path, user_id, scope, period_key):
        prompt_text, kb = build_simple_prompt(
            "Период уже закрыт. Правки недоступны.", f"v:{scope}:{period_key}"
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)
        return False
    if is_period_over(scope, period_key, today_date()):
        prompt_text, kb = build_simple_prompt(
            "Нельзя добавлять задачи в прошедший период.", f"v:{scope}:{period_key}"
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)
        return False
    return True


async def handle_start(
    message: Message,
    state: FSMContext,
    bot: Bot,
    db_path: str,
    scheduler: AsyncIOScheduler,
    admin_user_id: int | None,
) -> None:
    await ensure_user(db_path, message)
    await finalize_overdue_periods(db_path, message.from_user.id)
    user = await get_user(db_path, message.from_user.id)
    if user:
        await schedule_user_notifications(
            scheduler,
            bot,
            db_path,
            user["user_id"],
            user.get("notify_time") or "06:00",
            int(user.get("notify_enabled", 1)),
        )
    await state.clear()
    is_admin = admin_user_id is not None and message.from_user.id == admin_user_id
    await show_main_menu(bot, db_path, message.from_user.id, message.chat.id, is_admin)


async def handle_message_input(
    message: Message,
    state: FSMContext,
    bot: Bot,
    db_path: str,
    scheduler: AsyncIOScheduler,
    admin_user_id: int | None,
) -> None:
    user_id = message.from_user.id
    chat_id = message.chat.id
    data = await state.get_data()

    if await state.get_state() == InputStates.add_task.state:
        scope = data.get("scope")
        period_key = data.get("period_key")
        if await is_period_finalized(db_path, user_id, scope, period_key):
            await state.clear()
            prompt_text, kb = build_simple_prompt(
                "Период уже закрыт. Правки недоступны.",
                f"v:{scope}:{period_key}",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        if is_period_over(scope, period_key, today_date()):
            await state.clear()
            prompt_text, kb = build_simple_prompt(
                "Нельзя добавлять задачи в прошедший период.",
                f"v:{scope}:{period_key}",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        title, description = parse_title_description(message.text or "")
        if not title:
            prompt_text, kb = build_simple_prompt(
                "Название не может быть пустым. Отправьте название задачи.",
                f"v:{scope}:{period_key}",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        await create_task(db_path, user_id, scope, period_key, title, description)
        await state.clear()
        await show_plan(bot, db_path, user_id, chat_id, scope, period_key)

    elif await state.get_state() == InputStates.edit_task.state:
        task_id = data.get("task_id")
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await state.clear()
            await show_main_menu(bot, db_path, user_id, chat_id, user_id == admin_user_id)
            return
        if await is_period_finalized(db_path, user_id, task["scope"], task["period_key"]):
            await state.clear()
            prompt_text, kb = build_simple_prompt(
                "Период уже закрыт. Правки недоступны.",
                f"v:{task['scope']}:{task['period_key']}",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        if edit_policy(task["scope"], task["period_key"], today_date()) == "deny":
            await state.clear()
            prompt_text, kb = build_simple_prompt(
                edit_denied_message(task["scope"]),
                f"v:{task['scope']}:{task['period_key']}",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        title, description = parse_title_description(message.text or "")
        if not title:
            prompt_text, kb = build_simple_prompt(
                "Название не может быть пустым. Отправьте новое название задачи.",
                f"v:{task['scope']}:{task['period_key']}",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        await update_task(db_path, task_id, title, description)
        if task.get("is_locked"):
            recurring_id = task.get("recurring_id")
            if recurring_id:
                await update_recurring_task(db_path, recurring_id, title, description)
            else:
                recurring_id = await create_recurring_task(
                    db_path, user_id, task["scope"], title, description
                )
                await update_task_lock(db_path, task_id, 1, recurring_id)
        await state.clear()
        await show_plan(bot, db_path, user_id, chat_id, task["scope"], task["period_key"])

    elif await state.get_state() == InputStates.set_notify_time.state:
        notify_time = normalize_time_text(message.text or "")
        if not notify_time:
            prompt_text, kb = build_simple_prompt(
                "Некорректное время. Формат: HH:MM (например 06:30)",
                "st",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        await update_notify_settings(db_path, user_id, notify_time=notify_time)
        await log_history(
            db_path,
            user_id,
            "settings_update",
            {"notify_time": notify_time},
        )
        user = await get_user(db_path, user_id)
        if user:
            await schedule_user_notifications(
                scheduler,
                bot,
                db_path,
                user_id,
                user.get("notify_time") or notify_time,
                int(user.get("notify_enabled", 1)),
            )
        await state.clear()
        await show_settings(bot, db_path, user_id, chat_id)

    elif await state.get_state() == InputStates.edit_prayer_title.state:
        weekday = data.get("weekday")
        text = (message.text or "").strip()
        if text == "-":
            await set_prayer_entry(db_path, user_id, weekday, None, None)
            await log_history(
                db_path,
                user_id,
                "prayer_update",
                {"weekday": weekday, "title": None, "description": None},
            )
            await state.clear()
            await show_prayer_plan(bot, db_path, user_id, chat_id)
            return
        if not text:
            prompt_text, kb = build_simple_prompt(
                "Название не может быть пустым. Отправьте название или '-' чтобы очистить.",
                "pr",
            )
            await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)
            return
        await state.set_state(InputStates.edit_prayer_desc)
        await state.update_data(prayer_title=text)
        prompt_text, kb = build_simple_prompt(
            "Введите описание (опционально). Отправьте '-' чтобы оставить пустым.",
            "pr",
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb)

    elif await state.get_state() == InputStates.edit_prayer_desc.state:
        weekday = data.get("weekday")
        title = data.get("prayer_title")
        text = (message.text or "").strip()
        description = None if text in ("", "-") else text
        await set_prayer_entry(db_path, user_id, weekday, title, description)
        await log_history(
            db_path,
            user_id,
            "prayer_update",
            {"weekday": weekday, "title": title, "description": description},
        )
        await state.clear()
        await show_prayer_plan(bot, db_path, user_id, chat_id)

    else:
        await state.clear()
        is_admin = admin_user_id is not None and user_id == admin_user_id
        await show_main_menu(bot, db_path, user_id, chat_id, is_admin)

    try:
        await message.delete()
    except TelegramBadRequest:
        pass


async def handle_callback(
    callback: CallbackQuery,
    state: FSMContext,
    bot: Bot,
    db_path: str,
    scheduler: AsyncIOScheduler,
    admin_user_id: int | None,
) -> None:
    await ensure_user(db_path, callback)
    await state.clear()
    user_id = callback.from_user.id
    chat_id = callback.message.chat.id
    message_id = callback.message.message_id

    action, args = parse_callback(callback.data or "")
    force = False
    if action == "cf" and args:
        force = True
        action = args[0]
        args = args[1:]

    if action == "m":
        is_admin = admin_user_id is not None and user_id == admin_user_id
        await show_main_menu(bot, db_path, user_id, chat_id, is_admin, message_id)

    elif action == "p":
        today = today_date().strftime("%Y-%m-%d")
        await show_plan(bot, db_path, user_id, chat_id, "day", today, message_id)

    elif action == "v" and len(args) == 2:
        scope, period_key = args
        await show_plan(bot, db_path, user_id, chat_id, scope, period_key, message_id)

    elif action == "nav" and len(args) == 3:
        scope, period_key, direction = args
        delta = int(direction)
        from utils import add_period

        new_period = add_period(scope, period_key, delta)
        await show_plan(bot, db_path, user_id, chat_id, scope, new_period, message_id)

    elif action in ("t", "td") and len(args) == 1:
        task_id = int(args[0])
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await callback.answer("Задача не найдена", show_alert=False)
            return
        text, kb = build_task_detail_view(task)
        await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "tdg" and len(args) == 1:
        task_id = int(args[0])
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await callback.answer("Задача не найдена", show_alert=False)
            return
        if await is_period_finalized(db_path, user_id, task.get("scope"), task.get("period_key")):
            await callback.answer("Период закрыт для изменений", show_alert=False)
            return
        if not is_current_period(task.get("scope"), task.get("period_key"), today_date()):
            await callback.answer("Отметка доступна только в текущем периоде", show_alert=False)
            return
        new_state = 0 if task.get("is_done") else 1
        await toggle_task_done(db_path, task_id, new_state)
        task = await get_task(db_path, task_id)
        if task:
            text, kb = build_task_detail_view(task)
            await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "a" and len(args) == 2:
        scope, period_key = args
        allowed = await ensure_add_allowed(
            bot, db_path, user_id, chat_id, scope, period_key, message_id
        )
        if not allowed:
            return
        await state.set_state(InputStates.add_task)
        await state.update_data(scope=scope, period_key=period_key)
        prompt_text, kb = build_simple_prompt(
            "Введите задачу. Формат: строка 1 - название, строка 2 - описание (опционально).",
            f"v:{scope}:{period_key}",
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)

    elif action == "el" and len(args) == 2:
        scope, period_key = args
        allowed = await ensure_edit_allowed(
            bot, db_path, user_id, chat_id, scope, period_key, "el", message_id, force=force
        )
        if not allowed:
            return
        tasks = await list_tasks(db_path, user_id, scope, period_key)
        text, kb = build_task_select_view("Редактирование", tasks, "e", f"v:{scope}:{period_key}")
        await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "e" and len(args) == 1:
        task_id = int(args[0])
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await callback.answer("Задача не найдена", show_alert=False)
            return
        allowed = await ensure_edit_allowed(
            bot,
            db_path,
            user_id,
            chat_id,
            task["scope"],
            task["period_key"],
            "e",
            message_id,
            force=force,
            confirm_data=f"cf:e:{task_id}",
            back_data=f"v:{task['scope']}:{task['period_key']}",
        )
        if not allowed:
            return
        await state.set_state(InputStates.edit_task)
        await state.update_data(task_id=task_id)
        title = truncate(task.get("title", ""), 80)
        prompt_text, kb = build_simple_prompt(
            f"Текущее название: {title}\nОтправьте новое название и описание.",
            f"v:{task['scope']}:{task['period_key']}",
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)

    elif action == "dl" and len(args) == 2:
        scope, period_key = args
        allowed = await ensure_edit_allowed(
            bot, db_path, user_id, chat_id, scope, period_key, "dl", message_id, force=force
        )
        if not allowed:
            return
        tasks = await list_tasks(db_path, user_id, scope, period_key)
        text, kb = build_task_select_view("Удаление", tasks, "d", f"v:{scope}:{period_key}")
        await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "lk" and len(args) == 2:
        scope, period_key = args
        allowed = await ensure_edit_allowed(
            bot, db_path, user_id, chat_id, scope, period_key, "lk", message_id, force=force
        )
        if not allowed:
            return
        tasks = await list_tasks(db_path, user_id, scope, period_key)
        text, kb = build_lock_select_view(tasks, f"v:{scope}:{period_key}")
        await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "d" and len(args) == 1:
        task_id = int(args[0])
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await callback.answer("Задача не найдена", show_alert=False)
            return
        allowed = await ensure_edit_allowed(
            bot,
            db_path,
            user_id,
            chat_id,
            task["scope"],
            task["period_key"],
            "d",
            message_id,
            force=force,
            confirm_data=f"cf:d:{task_id}",
            back_data=f"v:{task['scope']}:{task['period_key']}",
        )
        if not allowed:
            return
        await delete_task(db_path, task_id)
        if task.get("is_locked") and task.get("recurring_id"):
            await delete_recurring_task(db_path, task["recurring_id"])
        await show_plan(bot, db_path, user_id, chat_id, task["scope"], task["period_key"], message_id)

    elif action == "lkc" and len(args) == 1:
        task_id = int(args[0])
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await callback.answer("Задача не найдена", show_alert=False)
            return
        if task.get("source") in ("prayer", "prayer_default"):
            await callback.answer("Молитвенный план уже цикличен", show_alert=False)
            return
        allowed = await ensure_edit_allowed(
            bot,
            db_path,
            user_id,
            chat_id,
            task["scope"],
            task["period_key"],
            "lkc",
            message_id,
            force=force,
            confirm_data=f"cf:lkc:{task_id}",
            back_data=f"v:{task['scope']}:{task['period_key']}",
        )
        if not allowed:
            return
        if task.get("is_locked"):
            if task.get("recurring_id"):
                await delete_recurring_task(db_path, task["recurring_id"])
            await update_task_lock(db_path, task_id, 0, None)
        else:
            recurring_id = task.get("recurring_id")
            if recurring_id:
                await update_recurring_task(
                    db_path, recurring_id, task.get("title", ""), task.get("description")
                )
            else:
                recurring_id = await create_recurring_task(
                    db_path, user_id, task["scope"], task.get("title", ""), task.get("description")
                )
            await update_task_lock(db_path, task_id, 1, recurring_id)
        tasks = await list_tasks(db_path, user_id, task["scope"], task["period_key"])
        text, kb = build_lock_select_view(tasks, f"v:{task['scope']}:{task['period_key']}")
        await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "tdlk" and len(args) == 1:
        task_id = int(args[0])
        task = await get_task(db_path, task_id)
        if not task or task.get("user_id") != user_id:
            await callback.answer("Задача не найдена", show_alert=False)
            return
        if task.get("source") in ("prayer", "prayer_default"):
            await callback.answer("Молитвенный план уже цикличен", show_alert=False)
            return
        allowed = await ensure_edit_allowed(
            bot,
            db_path,
            user_id,
            chat_id,
            task["scope"],
            task["period_key"],
            "tdlk",
            message_id,
            force=force,
            confirm_data=f"cf:tdlk:{task_id}",
            back_data=f"td:{task_id}",
        )
        if not allowed:
            return
        if task.get("is_locked"):
            if task.get("recurring_id"):
                await delete_recurring_task(db_path, task["recurring_id"])
            await update_task_lock(db_path, task_id, 0, None)
        else:
            recurring_id = task.get("recurring_id")
            if recurring_id:
                await update_recurring_task(
                    db_path, recurring_id, task.get("title", ""), task.get("description")
                )
            else:
                recurring_id = await create_recurring_task(
                    db_path, user_id, task["scope"], task.get("title", ""), task.get("description")
                )
            await update_task_lock(db_path, task_id, 1, recurring_id)
        task = await get_task(db_path, task_id)
        if task:
            text, kb = build_task_detail_view(task)
            await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "pr":
        await show_prayer_plan(bot, db_path, user_id, chat_id, message_id)

    elif action == "help":
        text, kb = build_help_view()
        await edit_or_send(bot, db_path, user_id, chat_id, text, kb, message_id)

    elif action == "pw" and len(args) == 1:
        weekday = int(args[0])
        await state.set_state(InputStates.edit_prayer_title)
        await state.update_data(weekday=weekday)
        prompt_text, kb = build_simple_prompt(
            "Введите название молитвенной нужды. '-' чтобы очистить.",
            "pr",
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)

    elif action == "st":
        await show_settings(bot, db_path, user_id, chat_id, message_id)

    elif action == "nt":
        await state.set_state(InputStates.set_notify_time)
        prompt_text, kb = build_simple_prompt(
            "Введите время уведомлений в формате HH:MM (например 06:30)",
            "st",
        )
        await edit_or_send(bot, db_path, user_id, chat_id, prompt_text, kb, message_id)

    elif action == "ntg":
        user = await get_user(db_path, user_id)
        if not user:
            return
        new_value = 0 if user.get("notify_enabled") else 1
        await update_notify_settings(db_path, user_id, notify_enabled=new_value)
        await log_history(db_path, user_id, "settings_update", {"notify_enabled": new_value})
        await schedule_user_notifications(
            scheduler,
            bot,
            db_path,
            user_id,
            user.get("notify_time") or "06:00",
            new_value,
        )
        await show_settings(bot, db_path, user_id, chat_id, message_id)

    elif action == "h":
        await show_history(bot, db_path, user_id, chat_id, message_id)

    elif action == "ex":
        filename = await export_to_excel(db_path, user_id)
        await log_history(db_path, user_id, "export", {"filename": filename})
        await bot.send_document(chat_id, FSInputFile(filename))

    elif action == "adm":
        if admin_user_id is None or user_id != admin_user_id:
            await callback.answer("Недостаточно прав", show_alert=False)
        else:
            await show_admin_view(bot, db_path, user_id, chat_id, message_id)

    await callback.answer()


async def main() -> None:
    import os

    config = load_config()
    db_dir = os.path.dirname(config.db_path)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    await init_db(config.db_path)

    bot = Bot(token=config.bot_token)
    await bot.delete_webhook(drop_pending_updates=True)
    storage = MemoryStorage()
    dispatcher = Dispatcher(storage=storage)
    scheduler = AsyncIOScheduler()

    @dispatcher.message(CommandStart())
    async def start_handler(message: Message, state: FSMContext) -> None:
        await handle_start(message, state, bot, config.db_path, scheduler, config.admin_user_id)

    @dispatcher.message(Command("menu"))
    async def menu_handler(message: Message, state: FSMContext) -> None:
        await handle_start(message, state, bot, config.db_path, scheduler, config.admin_user_id)

    @dispatcher.message(Command("help"))
    async def help_handler(message: Message) -> None:
        text, kb = build_help_view()
        await edit_or_send(bot, config.db_path, message.from_user.id, message.chat.id, text, kb)

    @dispatcher.message(InputStates.add_task)
    @dispatcher.message(InputStates.edit_task)
    @dispatcher.message(InputStates.set_notify_time)
    @dispatcher.message(InputStates.edit_prayer_title)
    @dispatcher.message(InputStates.edit_prayer_desc)
    async def input_handler(message: Message, state: FSMContext) -> None:
        await handle_message_input(message, state, bot, config.db_path, scheduler, config.admin_user_id)

    @dispatcher.callback_query()
    async def callback_handler(callback: CallbackQuery, state: FSMContext) -> None:
        await handle_callback(callback, state, bot, config.db_path, scheduler, config.admin_user_id)

    scheduler.start()
    await schedule_all_users(scheduler, bot, config.db_path)

    try:
        await dispatcher.start_polling(bot)
    finally:
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
